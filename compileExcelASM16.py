#COMPILER FOR EXCEL-ASM16
#USE:   py compileExcelASM16.py [program.s] ROM.xlsx

import sys
import os
import time
import math
from openpyxl import load_workbook

compiled = False;
filePath = ""
spreadsheet = ""
startTime = 0
data = []
program = []
output = []

labelOpen = False
labelToUse = ""
RED = '\033[91m';
ENDCOLOR = '\033[0m';

def integerError(lineNumber):
    print(RED + "\tInteger outside of expected range, line: " + str(lineNumber)+ ENDCOLOR)
    compileResults()
    
def syntaxError(lineNumber):
    print(RED + "\tSyntax Error, line: " + str(lineNumber)+ ENDCOLOR)
    compileResults()  
    
def labelError(lineNumber):
    print(RED + "\tDouble label detected, line " + str(lineNumber) + ENDCOLOR)
    compileResults()
    
def referenceNotFoundError(labelName):
    print(RED + "\tReference to variable or label not found, " + str(labelName) + ENDCOLOR)
    compileResults()

def unrecognizedError(lineNumber):
    print(RED + "\tUnrecognized Instruction, line " + str(lineNumber) + ENDCOLOR)
    compileResults()

def varSequenceError(lineNumber):
    print(RED + "\tVariables must be defined before program code, line " + str(lineNumber) + ENDCOLOR)
    compileResults()

def varUseError(varName):
    print(RED + "\tVariable cannot be used like label, var: " + str(lineNumber) + ENDCOLOR)
    compileResults()

def orgError(lineNumber):
    print(RED + "\tProgram Count exceeds target address, line " + str(lineNumber) + ENDCOLOR)
    compileResults()
    
def incResourceError(resourceName, lineNumber):
    print(RED + "\tResource " + resourceName + " could not be found, line " + str(lineNumber) + ENDCOLOR)
    compileResults()    

def lengthError(exceededWords):
    print(RED + "\tProgram length exceeds available RAM by " + str(exceededWords) + " words" + ENDCOLOR)
    compileResults()

def ROMbookError():
    print(RED + "\tCould not save to specified workbook, make sure the file is closed and try again" + ENDCOLOR)
    exit()
    
def createLine(label, operations):
    return [label, operations]

def getCurrentAddress():
    address = len(data)
    for operations in program:
        address = address + len(operations[1])
    return address

def getLocationOfLabel(labelName):
    location = 0
    for var in data:
        if (labelName == var[0]):
            varUseError(labelName)
        location = location + 1
    for operations in program:
        if (labelName == operations[0]):
            return location
        location = location + len(operations[1])
    return -1

def getVarIndex(varName):
    i = 0
    for var in data:
        if (varName == var[0]):
            return i
        i = i + 1
    return -1    

def includeBIN(fileName):
    with open(fileName, "rb") as incFile:
        cycle = False
        lastValue = 0
        while (word := incFile.read(1)):
            value = int.from_bytes(word, "big")
            if (cycle):
                value = (lastValue * 256) + value
                if (value >= pow(2, 16)):
                    input()
                program.append(createLine("", [value]))
                cycle = False
            else:
                lastValue = value
                cycle = True
        if (cycle): #catching last single byte value
            program.append(createLine("", [(lastValue * 256)]))
            
    return

def parseNumber(numberString, lineNumber):
    prefix = numberString[0]
    numberString = numberString[1:]
    result = 0
    if (prefix == "$" or prefix == "@"):
        result = int(numberString, 16)
        if (result > 65535):
            integerError(lineNumber)
    elif (prefix == "#"):
        result = int(numberString)
        if (result > 65535):
            integerError(lineNumber)
    elif (prefix == "R" and numberString.isdigit() and int(numberString) <= 15):
        result = int(numberString)
        if (result > 15):
            integerError(lineNumber)
    else:   #is a label
        result = getLocationOfLabel(prefix + numberString)
        if (result == -1 and lineNumber == -1): #second time around
            referenceNotFoundError(prefix + numberString)
        elif (result == -1):
            return "LABEL-" + (prefix + numberString)            
        if (result > 65535):
            integerError(lineNumber)
    if (result < 0):
        integerError(lineNumber)
    return result

def encode(line, lineNumber):
    #convert to list of integers
    opcode = line[0]
    operand0 = 0
    operand1 = 0
    twoWord = False
    #check instruction format:
    if (opcode == "JMP"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        twoWord = True
        operand0 = int("0000", 16)
        operand1 = parseNumber(line[1], lineNumber)
    elif (opcode == "JEQ"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        twoWord = True
        operand0 = int("0100", 16)
        operand1 = parseNumber(line[1], lineNumber)
    elif (opcode == "JLT"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        twoWord = True
        operand0 = int("0200", 16)
        operand1 = parseNumber(line[1], lineNumber)
    elif (opcode == "JGE"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        twoWord = True
        operand0 = int("0300", 16)
        operand1 = parseNumber(line[1], lineNumber)
    elif (opcode == "LOAD"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        varVal = getVarIndex(line[2])    #check if references variable
        if (line[1][0] == "R" and not(str(varVal)[0] == "-")):
            twoWord = True
            operand0 = int("0400", 16) + (parseNumber(line[1], lineNumber) * 16)
            operand1 = varVal
        elif (line[1][0] == line[2][0] and line[1][0] == "R"):
            operand0 = int("1900", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        elif (line[1][0] == "R" and line[2][0] == "@"):
            twoWord = True
            operand0 = int("0400", 16) + (parseNumber(line[1], lineNumber) * 16)
            operand1 = parseNumber(line[2], lineNumber)
        elif (line[1][0] == "R" and (line[2][0] == "$" or line[2][0] == "#")):
            twoWord = True
            operand0 = int("0500", 16) + (parseNumber(line[1], lineNumber) * 16)
            operand1 = parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "STORE"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        if (line[1][0] == line[2][0] and line[1][0] == "R"):
            operand0 = int("0700", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        elif (line[1][0] == "R" and line[2][0] == "@"):
            twoWord = True
            operand0 = int("0600", 16) + (parseNumber(line[1], lineNumber) * 16)
            operand1 = parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "TRAN"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        if (line[1][0] == line[2][0] and line[1][0] == "R"):
            operand0 = int("0800", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "ADD"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        if (line[1][0] == line[2][0] and line[1][0] == "R"):
            operand0 = int("0900", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "SUB"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        if (line[1][0] == line[2][0] and line[1][0] == "R"):
            operand0 = int("0A00", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "MULT"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        if (line[1][0] == line[2][0] and line[1][0] == "R"):
            operand0 = int("0B00", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "DIV"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        if (line[1][0] == line[2][0] and line[1][0] == "R"):
            operand0 = int("0C00", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "INC"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        if (line[1][0] == "R"):
            operand0 = int("0D00", 16) + (parseNumber(line[1], lineNumber) * 16)
        else:
            syntaxError(lineNumber)
    elif (opcode == "DEC"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        if (line[1][0] == "R"):
            operand0 = int("0E00", 16) + (parseNumber(line[1], lineNumber) * 16)
        else:
            syntaxError(lineNumber)
    elif (opcode == "AND"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        if (line[1][0] == line[2][0] and line[1][0] == "R"):
            operand0 = int("0F00", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "OR"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        if (line[1][0] == line[2][0] and line[1][0] == "R"):
            operand0 = int("1000", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "XOR"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        if (line[1][0] == line[2][0] and line[1][0] == "R"):
            operand0 = int("1100", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "NOT"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        if (line[1][0] == "R"):
            operand0 = int("1200", 16) + (parseNumber(line[1], lineNumber) * 16)
        else:
            syntaxError(lineNumber)
    elif (opcode == "ROL"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        if (not(line[1][0]) == line[2][0] and line[1][0] == "R"):
            operand0 = int("1300", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "ROR"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        if (not(line[1][0] == line[2][0]) and line[1][0] == "R"):
            operand0 = int("1400", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "CMP"):
        if (not(len(line) == 3)):
            syntaxError(lineNumber)
        if (line[1][0] == line[2][0] and line[1][0] == "R"):
            operand0 = int("1500", 16) + (parseNumber(line[1], lineNumber) * 16) + parseNumber(line[2], lineNumber)
        else:
            syntaxError(lineNumber)
    elif (opcode == "CLC"):
        if (not(len(line) == 1)):
            syntaxError(lineNumber)
        operand0 = int("1600", 16)
    elif (opcode == "STC"):
        if (not(len(line) == 1)):
            syntaxError(lineNumber)
        operand0 = int("1700", 16)
    elif (opcode == "NOP"):
        if (not(len(line) == 1)):
            syntaxError(lineNumber)
        operand0 = int("1800", 16)
    elif (len(line) == 3 and line[1] == "="):  #variables
        if (len(program) > 0):
            varSequenceError(lineNumber)
        data.append(createLine(line[0], line[2]))
        return None
    elif (len(line) == 1 and line[0] == ".DATA"):
        return None
    elif (len(line) == 1 and line[0] == ".CODE"):
        #ADD JUMP TO START OF PROGRAM
        data.insert(0, createLine("", len(data) + 2))    #number of words in data section
        data.insert(0, createLine("", int("0000", 16)))   #JMP
        return None
    elif (opcode == "ORG"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        targetAddress = parseNumber(line[1], lineNumber)
        currentAddress = getCurrentAddress()
        if (currentAddress > targetAddress):
            orgError(lineNumber)
        while(currentAddress < targetAddress):
            program.append(createLine("", [0]))
            currentAddress = currentAddress + 1
        return None
    elif (opcode == ".INC"):
        if (not(len(line) == 2)):
            syntaxError(lineNumber)
        line[1] = line[1].replace("\"", "")
        line[1] = line[1].replace("\'", "")
        if (not(os.path.isfile(line[1]))):
            incResourceError(line[1], lineNumber)
        includeBIN(line[1])
        return None
    else:
        unrecognizedError(lineNumber)
    
    if (not twoWord):
        #print(str(hex(operand0)).upper())
        return [operand0]
    else:
        #print(str(hex(operand0)).upper() + ", " + str(hex(operand1)).upper())
        return [operand0, operand1]
        
def parseProgram():
    global output
    global compiled
    global data
    for value in data:
        if (len(value[0]) > 0):
            value[1] = parseNumber(value[1], 0)
        output.append(value[1])
    for operations in program:
        for value in operations[1]:
            output.append(value)
    if (len(output) > 65536):
        lengthError(len(output) - 65536)
    compiled = True
    return

def parseUnmarkedLabels():
    pLine = 0
    for operations in program:
        valLine = 0
        for val in operations[1]:
            if ("LABEL" in str(val)):
                program[pLine][1][valLine] = parseNumber(val[6:], -1)
            valLine = valLine + 1
        pLine = pLine + 1
    return

def compileASM(filepath):
    file = open(filepath, "r")
    lineNumber = 1  #file line number for specifying errors
    for line in file:
        #print(line)
        line = line.upper()
        line = line.split(";")  #getting rid of comments
        line[0] = line[0].replace("\n", "") #removing return line
        line[0] = line[0].replace("\r", "")
        line[0] = line[0].strip()
        #print(line)
        if (len(line[0]) > 0):
            parseLine(line[0], lineNumber)
        lineNumber = lineNumber + 1
    parseUnmarkedLabels()    
    parseProgram()
    compileResults()

def parseLine(line, lineNumber):
    global labelOpen
    global labelToUse
    labelLine = line.split(":");
    label = labelLine[0]
    if (":" in line and len(labelLine[1]) <= 1):
        if (labelOpen):
            labelError(lineNumber)
        labelToUse = label  #add a label with no operations to program
        labelOpen = True
        return
    elif (":" not in line):
        if (labelOpen):
            label = labelToUse
            labelOpen = False
        else:
            label = ""
    else:
        if (labelOpen):
            labelError(lineNumber)
        line = labelLine[1].strip()
    line = line.split(" ")
    #print(str(lineNumber) + "\t" + str(label) + "\t" + str(line))
    operations = encode(line, lineNumber)
    if (not(operations == None)):
        program.append(createLine(label, operations))
    return

def sendToSpreadsheet():
    #load excel file
    workbook = load_workbook(filename = spreadsheet)     
    #open workbook
    sheet = workbook.active
    try:
        i = 0
        while (i < pow(2, 16)):
            if (i < len(output)):
                sheet.cell(row = math.floor(i / 256) + 1, column = (i % 256) + 1, value = output[i])
            else:
                sheet.cell(row = math.floor(i / 256) + 1, column = (i % 256) + 1, value = 0)
            i = i + 1
        #save the file
        workbook.save(filename = spreadsheet)
    except:
        ROMbookError()
    return

def compileResults():
    if (not(compiled)):
        print(RED + "\tProgram could not be compiled" + ENDCOLOR)
    else:
        print("\tProgram compiled Successfully")
        #print(output)
    print("\tProgram length in words: " + str(getCurrentAddress()))
    print("\tWriting to spreadsheet ROM...")
    sendToSpreadsheet()
    print("\tFinished in " + str(time.time()-startTime)[:6] + "s")
    exit()

if __name__ == "__main__":
    startTime = time.time()
    os.system('color')
    print("\tStarting operation...")
    
    if (len(sys.argv) == 3):
        filePath = sys.argv[1]
        spreadsheet = sys.argv[2]
    elif (len(sys.argv) == 1):
        print(RED + "\tInsufficent arguments, no ASM file specified" + ENDCOLOR)
        compileResults()
    elif (len(sys.argv) == 2):
        print(RED + "\tInsufficent arguments, no target spreadsheet specified" + ENDCOLOR)
        compileResults()
    else:
        print(RED + "\tArguments too many" + ENDCOLOR)
        compileResults()
    
    if (not(os.path.isfile(filePath))):
        print(RED + "\tFile " + filePath + " not found" + ENDCOLOR)
        compileResults()
    compileASM(filePath)